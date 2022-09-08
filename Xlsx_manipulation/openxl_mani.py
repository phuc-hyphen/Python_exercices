from openpyxl import load_workbook
import warnings
import string

# # save the workbook in the dataframe
# create new workbook
# workbook = Workbook()
# sheet = workbook.active




#  read_excel
# ignoring warning
warnings.simplefilter(action='ignore', category=UserWarning)

workbook = load_workbook(filename="/home/huu-phuc-le/NDF_Projets_V2.xlsx")
# to see all the sheets you have available to work with.
# print(workbook.sheetnames)

# # selects sheets
first_sheet = workbook.active  # the first available sheet

sheet_data2 = workbook["Lignes à requalifier"]  # with the sheet names
workbook.get_sheet_by_name("NdF")  # deprecated
# Accessing to a  cell
sheet_data2["A1"]
data = sheet_data2.cell(row=10, column=6)
# Get cell data
sheet_data2["A1"].value
data = sheet_data2.cell(row=10, column=6).value

# set data to a cell -> data will not be saved until rewrite the file
sheet_data2["A1"] = "Carrenec"

# get all values in excel
for row in sheet_data2.values:
    for value in row:
        print(value)

for row in sheet_data2.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    for cell in row:
        print(cell)

# Create a dictionary of column names





dict = build_col_dict(sheet_data2)
for row_cells in sheet_data2.iter_rows(min_row=1, max_row=4):
    print(row_cells[dict['MOIS']].value)


# Accessing many cells
cell_range = sheet_data2['A1:C3']

# for row in sheet_data2.iter_rows(min_row=1, max_col=3, max_row=2, values_only=False):
#     for cell in row:
#         print(cell)
# for col in sheet_data2.iter_cols(min_row=1, max_col=3, max_row=2):
#     for cell in col:
#         print(cell)
print(cell_range)
# Accessing to Ranges of rows or columns
col_range = sheet_data2['A']
row_range = sheet_data2[10]
# print(sheet_data2.rows)

# save the workbook
workbook.save("/home/huu-phuc-le/NDF Projets.xlsx")
