from my_xlsx_utils import *
from openpyxl import Workbook
# ignoring warning
warnings.simplefilter(action='ignore', category=UserWarning)

NdF_data = load_workbook("xlsx_manipulation/NdF_Data.xlsx")

NdF_data_sheet = NdF_data["NdF"]

NdF_group = Workbook()

NdF_group_sheet = NdF_group.active

add_col(NdF_group_sheet, "employee")
add_col(NdF_group_sheet, "Montant Total expected")

NdF_sheet_dict = build_col_dict(NdF_data_sheet)
NdF_group_dict = build_col_dict(NdF_group_sheet)


def add_row(sheet, name_row):
    pass


# group by name les montant
current = ""
sum = 0
cur_row = 1
for row in NdF_data_sheet.iter_rows():
    if row[NdF_sheet_dict["Worker"]].value != current and row[NdF_sheet_dict["Worker"]].value != "Worker":
        current = row[NdF_sheet_dict["Worker"]].value
        sum += float(row[NdF_sheet_dict["MNT"]].value)
    else:
        cur_row += 1
        NdF_group_sheet.insert_rows(cur_row, 1)
        NdF_group_sheet.cell(
            column=NdF_group_dict["Montant Total expected"], row=cur_row).value = sum
        sum = 0

NdF_group.save("xlsx_manipulation/NdF_Data_grouped.xlsx")
