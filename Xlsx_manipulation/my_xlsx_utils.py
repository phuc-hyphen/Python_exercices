from openpyxl import load_workbook
import warnings
import string
import os
import pandas as pd


def add_col(sheet, name_col):
    nbr_col = sheet.max_column
    sheet.insert_cols(nbr_col + 1)
    location = string.ascii_uppercase[nbr_col] + "1"
    sheet[location] = name_col


def build_col_dict(sheet):
    ColNames = {}
    Current = 0
    for COL in sheet.iter_cols(1, sheet.max_column):
        ColNames[COL[0].value] = Current
        Current += 1
    return ColNames


def concat_files_excel_in_folder(path):
    dir_list = os.listdir(path)
    file = path + "/" + dir_list[0]
    df1 = pd.read_excel(file)

    for i in range(1, len(dir_list)):
        file_path = path + "/" + dir_list[i]
        df = pd.read_excel(file_path)
        df1 = pd.concat([df1, df])
    df1.to_excel(file)

concat_files_excel_in_folder("dossier_to_concat")